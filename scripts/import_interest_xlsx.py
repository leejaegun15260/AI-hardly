# -*- coding: utf-8 -*-
"""
'직업흥미유형별 군사특기' 엑셀 → 표준 데이터로 변환

입력: data/raw/직업흥미유형별_군사특기.xlsx
      시트 6개(현실형·탐구형·예술형·사회형·진취형·관습형)
      각 시트 열: 순번 | 군별 | 군사특기코드 | 군사특기 | 직업군

출력: data/build/specialty.json     군사특기 마스터 (군별+코드로 하나씩)
      data/build/interest_map.json  흥미유형 → 특기코드 목록

주의 — 코드 형식
  엑셀에 숫자로 저장돼 있어 그대로 읽으면 223.108 → 223.108(float) 이 됩니다.
  육군은 소수 3자리, 해·공군·해병은 자릿수가 제각각이라
  '엑셀 화면에 보이는 그대로'를 문자열로 복원해서 씁니다. (예: 17.01 → "17.01")
  나중에 병무청 API 코드와 맞출 때를 대비해 육군은 3자리로 채운 code_norm 도 같이 넣습니다.
"""
import json, pathlib, collections
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC  = ROOT / "data" / "raw" / "직업흥미유형별_군사특기.xlsx"
OUT  = ROOT / "data" / "build"

# 파일의 시트 이름 → RIASEC 코드
TYPE_CODE = {"현실형":"R","탐구형":"I","예술형":"A","사회형":"S","진취형":"E","관습형":"C"}


def code_text(v):
    """엑셀에 보이는 그대로의 코드 문자열로 되돌립니다."""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float):
        return ("%f" % v).rstrip("0").rstrip(".")
    return str(v)


def code_norm(branch, text):
    """육군 코드를 소수 3자리로 맞춥니다 (163.11 → 163.110). 다른 군은 그대로."""
    if branch != "육군" or "." not in text:
        return text
    head, tail = text.split(".", 1)
    return f"{head}.{tail.ljust(3, '0')}"


def main():
    if not SRC.exists():
        raise SystemExit(f"원본 파일이 없습니다: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True)
    unknown = [s for s in wb.sheetnames if s not in TYPE_CODE]
    if unknown:
        raise SystemExit(f"모르는 시트가 있습니다: {unknown}  (TYPE_CODE 를 확인하세요)")

    spec = collections.OrderedDict()   # key = "육군-162.104"
    rows_read = 0

    for ws in wb.worksheets:
        itype = TYPE_CODE[ws.title]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            _, branch, raw_code, name, job = row
            rows_read += 1
            text = code_text(raw_code)
            key  = f"{branch}-{text}"
            s = spec.setdefault(key, {
                "specialty_code": key,
                "branch": branch,
                "code": text,
                "code_norm": code_norm(branch, text),
                "specialty_name": str(name).strip(),
                "interest_types": [],
                "interest_labels": [],
                "job_groups": [],
                # 아래 항목은 병무청 오픈API(모집병 군사특기 정보)에서 채울 자리입니다.
                "field": None, "duty": None, "physical_grade": None, "interview": None,
                "source": "직업흥미유형별 군사특기 (병무청 제공 엑셀)",
            })
            if s["specialty_name"] != str(name).strip():
                print(f"  [경고] 같은 코드에 다른 이름: {key} "
                      f"'{s['specialty_name']}' vs '{name}'")
            if itype not in s["interest_types"]:
                s["interest_types"].append(itype)
                s["interest_labels"].append(ws.title)
            if job and str(job).strip() not in s["job_groups"]:
                s["job_groups"].append(str(job).strip())

    specialties = list(spec.values())
    for s in specialties:
        s["interest_types"].sort()
        s["interest_labels"].sort()
        s["job_groups"].sort()

    imap = {c: [] for c in TYPE_CODE.values()}
    for s in specialties:
        for t in s["interest_types"]:
            imap[t].append(s["specialty_code"])

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT/"specialty.json").write_text(
        json.dumps(specialties, ensure_ascii=False, indent=1), encoding="utf-8")
    (OUT/"interest_map.json").write_text(
        json.dumps(imap, ensure_ascii=False, indent=1), encoding="utf-8")

    by_branch = collections.Counter(s["branch"] for s in specialties)
    jobs = {j for s in specialties for j in s["job_groups"]}
    print(f"읽은 행 {rows_read}건 → 군사특기 {len(specialties)}개")
    print("  군별:", dict(by_branch))
    print("  흥미유형별:", {k: len(v) for k, v in imap.items()})
    print(f"  직업군 {len(jobs)}종")
    print(f"저장: {OUT/'specialty.json'}, {OUT/'interest_map.json'}")


if __name__ == "__main__":
    main()
