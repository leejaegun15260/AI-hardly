# -*- coding: utf-8 -*-
"""
'모집병 커트라인' 엑셀 → 표준 데이터로 변환

입력: data/raw/모집병_커트라인_25년8회차-26년8회차.xlsx
      시트 4개 (해군 · 공군 · 해병대 · 육군)
      1행: 회차 이름 (25년 8회차 … 26년 8회차)
      2행: 모집 월
      3행부터 자료
        육군      A열 특기코드 · B열 특기명 · C열~O열 회차별 커트라인
        그 외 군  A열 모집 구분 · B열 특기명 · C열~O열 회차별 커트라인

출력: data/build/cutoff.json   회차별 커트라인 (실제 자료)

주의 — 특기를 알아보는 법
  육군은 특기코드로 맞춥니다. 엑셀에 숫자로 저장돼 있어 그대로 읽으면
  163.11 과 163.110 이 달라 보이므로 소수 3자리로 맞춰서 이어 줍니다.

  해군은 특기명만으로는 구분이 안 됩니다. '조리'가 네 군데 나옵니다.
      전문기술          조리  →  해군-11.05
      복무지역(1함대)   조리  →  해군-A11.05
      복무지역(2함대)   조리  →  해군-B11.05
      작전사(진해,부산,제주) 조리 → 해군-C11.05
  그래서 A열의 모집 구분을 특기코드 앞글자(A·B·C)와 맞춰서 구분합니다.

  못 맞춘 줄도 버리지 않고 특기코드를 비워 둔 채 함께 담고,
  결과 파일의 '확인필요' 에 이유와 함께 적어 둡니다.

실행:  python3 scripts/import_cutoff_xlsx.py
"""
import json, pathlib, re, datetime
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC  = ROOT / "data" / "raw" / "모집병_커트라인_25년8회차-26년8회차.xlsx"
OUT  = ROOT / "data" / "build"

# 엑셀 시트 이름 → 특기 목록의 군 이름
SHEET_BRANCH = {"해군": "해군", "공군": "공군", "해병대": "해병", "육군": "육군"}

# 해군 A열의 모집 구분 → 특기코드 앞글자
#   특기 목록의 코드가 A11.05='(1함대)조리', B43='(2함대)기관', C43='(작전사)기관' 이므로
#   A=1함대, B=2함대, C=작전사 로 봅니다. (특기명 표기보다 코드가 일관적입니다)
AREA_PREFIX = {"복무지역(1함대)": "A", "복무지역(2함대)": "B", "작전사": "C"}

SCORE_COL_FROM, SCORE_COL_TO = 2, 15      # C열~O열 (0부터 셈)


def code_text(v):
    """엑셀에 보이는 그대로의 코드 문자열로 되돌립니다. (121.101 이 121.101 로 남도록)"""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, float):
        return ("%f" % v).rstrip("0").rstrip(".")
    return str(v)


def norm3(code):
    """소수점 아래를 세 자리로 맞춥니다. 163.11 → 163.110"""
    code = str(code).strip()
    # 끝에 점이 붙은 코드는 지우면 안 됩니다.
    #   '163.110' = 그레이더운용,  '163.110.' = 도로포장장비운용 — 서로 다른 특기입니다.
    if code.endswith("."):
        return code
    if "." in code:
        head, tail = code.split(".", 1)
        code = head + "." + (tail + "000")[:3]
    return code


def simple(s):
    """이름을 맞출 때 쓰는 형태 — 띄어쓰기·괄호·가운뎃점 등을 뺍니다."""
    return re.sub(r"[\s()·・/\-]", "", str(s or ""))


def parse_period(label, month):
    """'25년 8회차' → (2025, 8).  못 읽으면 (None, None)"""
    m = re.match(r"\s*(\d{2})년\s*(\d+)회차", str(label or ""))
    if not m:
        return None, None
    return 2000 + int(m.group(1)), int(m.group(2))


def load_specialty():
    f = OUT / "specialty.json"
    if not f.exists():
        raise SystemExit(f"특기 목록이 없습니다: {f}\n먼저 scripts/import_interest_xlsx.py 를 실행하세요.")
    spec = json.loads(f.read_text(encoding="utf-8"))

    by_code = {}          # "육군-121.101"        → 특기
    by_name = {}          # ("해군","일반","")    → [특기]   (군, 이름, 코드앞글자)
    for s in spec:
        b, code = s["branch"], s["code"]
        by_code[f"{b}-{norm3(code)}"] = s
        prefix = code[0] if code[:1] in ("A", "B", "C") else ""
        # '(1함대)조리' 처럼 앞에 붙은 괄호는 떼고 이름만 씁니다
        base = re.sub(r"^\([^)]*\)", "", s["specialty_name"])
        by_name.setdefault((b, simple(base), prefix), []).append(s)
    return spec, by_code, by_name


def main():
    if not SRC.exists():
        raise SystemExit(f"원본 파일이 없습니다: {SRC}")

    spec, by_code, by_name = load_specialty()
    wb = openpyxl.load_workbook(SRC, data_only=True)

    unknown = [s for s in wb.sheetnames if s not in SHEET_BRANCH]
    if unknown:
        raise SystemExit(f"모르는 시트가 있습니다: {unknown}  (SHEET_BRANCH 를 확인하세요)")

    # ---------- 회차 목록 (모든 시트가 같아야 합니다) ----------
    periods = None
    for ws in wb.worksheets:
        head  = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        month = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        got = []
        for i in range(SCORE_COL_FROM, SCORE_COL_TO):
            label = head[i] if i < len(head) else None
            if not label:
                continue
            y, r = parse_period(label, month[i] if i < len(month) else None)
            got.append({"col": i, "label": str(label).strip(),
                        "short": f"{str(y)[2:]}-{r}" if y else str(label).strip(),
                        "month": str(month[i]).strip() if i < len(month) and month[i] else "",
                        "year": y, "round": r})
        if periods is None:
            periods = got
        elif [p["label"] for p in got] != [p["label"] for p in periods]:
            raise SystemExit(f"[{ws.title}] 시트의 회차가 다른 시트와 다릅니다.")
    for n, p in enumerate(periods, 1):
        p["seq"] = n

    # ---------- 자료 읽기 ----------
    rows, need_check = [], []
    seen = {}

    for sheet, branch in SHEET_BRANCH.items():
        ws = wb[sheet]
        group = ""
        for excel_row, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            code_cell = r[0] if len(r) > 0 else None
            name      = str(r[1]).strip() if len(r) > 1 and r[1] else ""
            if branch != "육군" and code_cell and str(code_cell).strip():
                group = str(code_cell).strip()          # 위아래로 합쳐진 칸이라 이어서 씁니다
            if not name:
                continue

            scores = []
            for p in periods:
                v = r[p["col"]] if p["col"] < len(r) else None
                scores.append(round(float(v), 1) if isinstance(v, (int, float)) else None)
            if not any(s is not None for s in scores):
                continue

            # ---- 어느 특기인지 찾기 ----
            hit, why = None, ""
            if branch == "육군":
                key = f"육군-{norm3(code_text(code_cell))}"
                hit = by_code.get(key)
                if not hit:
                    why = f"특기 목록에 없는 특기코드({code_text(code_cell)})"
            else:
                prefix = ""
                for gname, pre in AREA_PREFIX.items():
                    if group.startswith(gname[:6]):
                        prefix = pre
                        break
                cand = by_name.get((branch, simple(name), prefix), [])
                if len(cand) == 1:
                    hit = cand[0]
                elif len(cand) > 1:
                    why = f"같은 이름의 특기가 {len(cand)}개라 고를 수 없음"
                else:
                    why = f"특기 목록에 없는 이름(구분 '{group}')"

            row = {
                "branch": branch,
                "specialty_code": hit["specialty_code"] if hit else None,
                "name": name,
                "group": group if branch != "육군" else "",
                "code_raw": code_text(code_cell) if branch == "육군" else "",
                "scores": scores,
            }
            rows.append(row)
            if not hit:
                need_check.append({"군": branch, "구분": row["group"], "특기명": name,
                                   "특기코드": row["code_raw"], "이유": why,
                                   "엑셀_줄": f"{sheet} {excel_row}행"})
            else:
                seen.setdefault(hit["specialty_code"], []).append(f"{sheet} {excel_row}행")

    dup = {k: v for k, v in seen.items() if len(v) > 1}

    out = {
        "_설명": [
            "병무청 제공 엑셀에서 옮긴 회차별 합격 커트라인입니다. 실제 자료입니다.",
            "periods 는 회차 목록(오래된 것부터), rows[].scores 는 그 순서대로의 커트라인입니다.",
            "scores 의 null 은 그 회차에 모집이 없었거나 자료가 없다는 뜻입니다.",
            "specialty_code 가 null 인 줄은 특기 목록과 이어 붙이지 못한 것으로, 확인필요 에 적혀 있습니다.",
        ],
        "source": "병무청 제공 엑셀 (25년 8회차 ~ 26년 8회차)",
        "source_file": SRC.name,
        "imported_at": datetime.date.today().isoformat(),
        "periods": [{k: v for k, v in p.items() if k != "col"} for p in periods],
        "rows": rows,
        "확인필요": need_check,
        "특기가_겹치는_줄": dup,
    }
    (OUT / "cutoff.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    # ---------- 요약 ----------
    matched = [r for r in rows if r["specialty_code"]]
    n_scores = sum(1 for r in rows for s in r["scores"] if s is not None)
    print(f"회차 {len(periods)}개: {periods[0]['label']} ~ {periods[-1]['label']}")
    print(f"자료 줄 {len(rows)}개 · 커트라인 숫자 {n_scores}개")
    for b in ("육군", "해군", "공군", "해병"):
        got = [r for r in rows if r["branch"] == b]
        ok  = [r for r in got if r["specialty_code"]]
        print(f"  {b}  {len(got):3d}줄 중 {len(ok):3d}줄을 특기와 이어 붙였습니다")
    if need_check:
        print(f"\n이어 붙이지 못한 줄 {len(need_check)}개 — 확인이 필요합니다:")
        for c in need_check:
            print(f"  · {c['군']} {c['구분']} {c['특기명']} {c['특기코드']} — {c['이유']} ({c['엑셀_줄']})")
    if dup:
        print(f"\n한 특기에 두 줄 이상이 붙었습니다 — 확인이 필요합니다:")
        for k, v in dup.items():
            print(f"  · {k} ← {v}")
    print(f"\n생성: {OUT/'cutoff.json'}  ({(OUT/'cutoff.json').stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    main()
